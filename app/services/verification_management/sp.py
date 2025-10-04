from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from services.sw_utils.errors import SwTableError

# from .excel_utils import
from services.verification_management.excel_utils import *
from services.verification_management.matches import *

BASE_DIR_SP: str = 'Проверка СП'
TEMPLATE_DIR_SP_SW: str = r'C:\SWR-Библиотеки 2021\Мое\Справка\Для проги\Шаблон СП сборки.sldbomtbt'


def get_dir_sp(dir_project: str, number_project: str) -> str:
    dir_sp: Path = Path(dir_project).parent / BASE_DIR_SP
    if not dir_sp.exists():
        raise FileNotFoundError(f'Папка {dir_sp} не существует')

    for file in dir_sp.iterdir():
        if file.is_file():
            if number_project in file.stem:
                return str(file)
    else:
        raise FileNotFoundError(f'Специя с номером проекта {number_project} не найдена')


def save_sp_assem(sw_assem, dir_sp: str, number_project: str) -> str | None:
    sp_assem = sw_assem.Extension.InsertBomTable3(TEMPLATE_DIR_SP_SW, 0, 0, 1, 'По умолчанию', False, 0, False)
    if not sp_assem:
        raise SwTableError
    dir_sp_assem: Path = Path(dir_sp).parent / f'СП_{number_project}_SW.xlsx'
    sp_assem.SaveAsExcel(str(dir_sp_assem), False, False)
    sp_assem.BomFeature.GetFeature.Select2(False, 0)
    sw_assem.Extension.DeleteSelection2(1)
    return str(dir_sp_assem)


def verification_excel(ws_sp, ws_sp_assem) -> list[tuple]:
    sp1: list[tuple] = [row for row in get_cell(ws_sp, min_col=1, max_col=5, min_row=3, values_only=True)]
    sp2: list[tuple] = [row for row in get_cell(ws_sp_assem, min_row=2, values_only=True)]

    exact_matches, sp1, sp2 = get_exact_matches(sp1, sp2)

    fuzzy_matches, sp1, sp2 = get_fuzzy_matches(sp1, sp2)

    thresholds: list[int] = [70, 60, 50, 40, 30]
    last_matches: list[tuple] = []
    for t in thresholds:
        matches, sp1, sp2 = match_items(sp1, sp2, t)
        last_matches.extend(matches)

    last = zip_last(sp1, sp2)

    return [('-',), *exact_matches, ('-',), *fuzzy_matches, ('-',), *last_matches, ('-',), *last]


def create_verification_excel(dir_sp: str, dir_sp_assem: str):
    # wb_sp = load_workbook(dir_sp)
    # ws_sp = wb_sp.active
    #
    # wb_sp_assem = load_workbook(dir_sp_assem)
    # ws_sp_assem = wb_sp_assem.active

    wb_sp = load_workbook(r'C:\Users\Вадим\Desktop\Тестирование\8902\КД\Проверка СП\СП_8902.xlsx')
    ws_sp = wb_sp.active

    wb_sp_assem = load_workbook(r'C:\Users\Вадим\Desktop\Тестирование\8902\КД\Проверка СП\СП_8902_SW.xlsx')
    ws_sp_assem = wb_sp_assem.active

    matches: list[tuple] = verification_excel(ws_sp, ws_sp_assem)

    wb_compare = Workbook()
    ws_compare = wb_compare.active

    add_headers(ws_compare)
    [ws_compare.append(i) for i in matches]

    add_alignment(ws_compare)

    edit_width_ws(ws_compare)

    add_fill(ws_compare)

    add_border(ws_compare)

    # wb_compare.save(str(Path(dir_sp).parent / Path(dir_sp).name) + ' сравнение.xlsx')
    wb_compare.save(r'C:\Users\Вадим\Desktop\Тестирование\8902\КД\Проверка СП\СП_8902 сравнение.xlsx')


create_verification_excel('', '')
