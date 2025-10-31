from openpyxl.styles import PatternFill, Alignment, Side, Border


def create_style_excel(ws, matches: dict[str, list]) -> None:
    _edit_width_ws(ws)

    _add_headers(ws)
    _add_subheaders(ws, 'Совпадение на 100%', '548235')
    for rows in matches['exact_matches']:
        for row in rows:
            ws.append(row)
        ws.append(('',))

    _add_subheaders(ws, 'Совпадение на 90%', 'A9D08E')
    for rows in matches['fuzzy_matches']:
        for row in rows:
            ws.append(row)
        ws.append(('',))

    _add_subheaders(ws, 'Совпадение < 90%', 'C6E0B4')
    for rows in matches['last_matches']:
        for row in rows:
            ws.append(row)
        ws.append(('',))

    _add_fill_matches(ws)
    _add_alignment(ws)

    current_row: int = ws.max_row + 2
    _add_subheaders(ws, 'Оставшиеся позиции из Excel', 'E2EFDA')
    for row in matches['spec']:
        ws.append([row[0], row[1], row[2], '', row[3]])
    _add_fill_last_items(ws, current_row, 'FFF2CC')

    current_row: int = ws.max_row + 2
    _add_subheaders(ws, 'Оставшиеся позиции из сборки', 'EBF1DE')
    for row in matches['model_spec']:
        ws.append([row[0], row[1], row[2], '', row[3]])
    _add_fill_last_items(ws, current_row, 'BDD7EE')

    current_row: int = ws.max_row + 2
    _add_subheaders(ws, 'Фасонные части', 'EBF1DE')
    for row in matches['secondary_items']:
        ws.append([row[0], row[1], row[2], '', row[3]])
    _add_fill_last_items(ws, current_row, 'BDD7EE')

    _add_merge_empty_row_border(ws)


def _get_cell(
        ws,
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        values_only: bool = False
):
    return list(ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=values_only
    ))


def _add_headers(ws) -> None:
    headers: tuple = (
        'Артикул',
        'Кол-во',
        'Совп.',
        'Ед. Изм.',
        'Наименование',
    )
    ws.append(headers)


def _add_subheaders(ws, value: str, fill: str) -> None:
    cell = ws.cell(row=ws.max_row + 1, column=1, value=value)
    ws.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=5)
    _apply_fill(fill, (cell,))


def _add_alignment(ws):
    alignments = {
        'left': ['A', 'E'],
        'center': ['B', 'C', 'D'],
    }

    for align, cols in alignments.items():
        for col in cols:
            for cell in ws[col]:
                cell.alignment = Alignment(horizontal=align, vertical='center')


def _edit_width_ws(ws) -> None:
    widths = {
        'A': 30,
        'B': 7,
        'C': 7,
        'D': 8,
        'E': 200,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _add_fill_matches(ws) -> None:
    _apply_fill('FFC000', *_get_cell(ws, min_col=1, min_row=1, max_row=1))  # Заголовок
    counter: int = 1
    for row in _get_cell(ws, min_row=3):
        if row[4].value is None:
            counter = 1
            continue
        if counter == 2:
            _apply_fill('BDD7EE', row)  # СП сборки
            continue
        _apply_fill('FFF2CC', row)  # СП
        _add_merge_row(ws, row[2].row)
        if row[2].value == 0:
            _apply_fill('92D050', (row[2],))
        else:
            _apply_fill('FF0000', (row[2],))
        counter += 1


def _add_fill_last_items(ws, current_row: int, fill: str) -> None:
    for row in _get_cell(ws, min_row=current_row):
        _apply_fill(fill, row)


def _apply_fill(fill: str, row) -> None:
    for cell in row:
        cell.fill = _create_fill(fill)


def _create_fill(rgb: str):
    return PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')


def _add_merge_row(ws, row) -> None:
    ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=3)


def _add_merge_empty_row_border(ws) -> None:
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for row in _get_cell(ws):
        for cell in (cell for row in _get_cell(ws) for cell in row):
            cell.border = border
        if row[4].value is None:
            ws.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=5)


def _add_border(ws) -> None:
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in (cell for row in _get_cell(ws) for cell in row):
        cell.border = border
