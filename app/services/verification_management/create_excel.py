from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from .excel_utils import create_style_excel
from .matches import get_matches


def create_verification_excel(spec_dir: str, model_spec_dir: str):
    spec_wb = load_workbook(spec_dir)
    spec_ws = spec_wb.active

    model_spec_wb = load_workbook(model_spec_dir)
    model_spec_ws = model_spec_wb.active

    matches: dict[str, list] = get_matches(spec_ws, model_spec_ws)

    compare_wb = Workbook()
    compare_ws = compare_wb.active

    create_style_excel(compare_ws, matches)

    compare_wb.save(str(
        Path(spec_dir).parent / Path(spec_dir).stem) + ' сравнение.xlsx'
                    )
